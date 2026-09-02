#!/usr/bin/env python3
"""
compile_data.py — My Song Writer data compiler

Reads genre_chord_bank_v*.xlsx and emits song_data.js, a zero-dependency
JS data module to embed (or <script src=>) in the single-file app.

Usage:
    python compile_data.py [workbook.xlsx] [output.js] [--embed [index.html]]
    Defaults: genre_chord_bank_v9.xlsx -> song_data.js
    --embed also splices the generated module into index.html between the
    SONG_DATA_BEGIN / SONG_DATA_END markers (single-file rule).

Re-run whenever the workbook changes (e.g., after ear-checking pattern
mappings). Validation warnings print to stderr; they do not block output.

Requires: openpyxl  (pip install openpyxl)
"""

import json
import re
import sys
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

_args = [a for a in sys.argv[1:] if not a.startswith("--")]
EMBED = None
if "--embed" in sys.argv:
    i = sys.argv.index("--embed")
    EMBED = sys.argv[i + 1] if i + 1 < len(sys.argv) and not sys.argv[i + 1].startswith("--") and sys.argv[i + 1].endswith(".html") else "index.html"
    _args = [a for a in _args if a != EMBED]
WB_PATH = _args[0] if len(_args) > 0 else "genre_chord_bank_v9.xlsx"
OUT_PATH = _args[1] if len(_args) > 1 else "song_data.js"

warnings = []


def warn(msg):
    warnings.append(msg)


def clean(v):
    """Normalize a cell value: strip strings, keep numbers, None for blanks."""
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def rows_as_dicts(ws):
    """First row = headers; remaining rows -> list of dicts, skipping blank rows."""
    rows = list(ws.iter_rows(values_only=True))
    headers = [clean(h) for h in rows[0]]
    out = []
    for raw in rows[1:]:
        vals = [clean(v) for v in raw]
        if all(v is None for v in vals):
            continue
        out.append({h: v for h, v in zip(headers, vals) if h is not None})
    return out


# --- parsers -----------------------------------------------------------------

DASHES = "\u2013\u2014\u2212-"  # en dash, em dash, minus, hyphen


def parse_range(s):
    """'40–80' / '40-80' -> [40, 80]; 'any'/None -> None."""
    if s is None:
        return None
    s = str(s).strip()
    if s.lower() == "any":
        return None
    m = re.match(rf"^(\d+)\s*[{DASHES}]\s*(\d+)$", s)
    if m:
        return [int(m.group(1)), int(m.group(2))]
    if s.isdigit():
        n = int(s)
        return [n, n]
    warn(f"Unparsed range: {s!r}")
    return None


def parse_slots(s):
    """
    Duration group -> list of slot durations in ms (at the 60 BPM reference).
      '4000'              -> [4000]
      '2000 + 2000'       -> [2000, 2000]
      '500 × 8' / '500 x 8' -> [500]*8
      '1500 + 1500 + 1000'-> [1500, 1500, 1000]
      'any' / None        -> None (custom pattern)
    """
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return [int(s)]
    s = str(s).strip()
    if s.lower() == "any":
        return None
    # 'A then B ring' -> slots(A) + [B]
    m = re.match(r"^(.*?)\s+then\s+(\d+)\s*ring$", s, re.I)
    if m:
        head = parse_slots(m.group(1))
        return (head or []) + [int(m.group(2))]
    # '(a + b) x n' -> [a, b] * n
    m = re.match(r"^\((.*?)\)\s*[x\u00d7]\s*(\d+)$", s, re.I)
    if m:
        grp = parse_slots(m.group(1))
        return (grp or []) * int(m.group(2))
    m = re.match(r"^(\d+)\s*[x\u00d7]\s*(\d+)$", s, re.I)
    if m:
        return [int(m.group(1))] * int(m.group(2))
    parts = re.split(r"\s*\+\s*", s)
    if all(p.strip().isdigit() for p in parts):
        return [int(p) for p in parts]
    warn(f"Unparsed duration group: {s!r}")
    return None


def parse_bars(s):
    """'1–8' -> [1, 8]; '5' -> [5, 5]."""
    r = parse_range(s)
    if r is None:
        warn(f"Unparsed bar range: {s!r}")
    return r


def split_list(s, sep=";"):
    """'A; B; C' -> ['A','B','C']; '—'/None -> []."""
    if s is None:
        return []
    s = str(s).strip()
    if s in ("\u2014", "-", "\u2013", "none", "None", ""):
        return []
    return [p.strip() for p in s.split(sep) if p.strip()]


def pattern_common(row):
    """Fields shared by all pattern banks."""
    return {
        "name": row.get("Pattern"),
        "slots": parse_slots(row.get("Duration group (ms)")),
        "readsAs": row.get("Reads as"),
        "bpmRange": parse_range(row.get("BPM Range")),
        "slot": row.get("Slot"),
        "ratio": row.get("Ratio"),
        "sustain": row.get("Sustain"),
        "mute": row.get("Mute"),
        "custom": str(row.get("Pattern", "")).lower().startswith("custom"),
    }


# --- load --------------------------------------------------------------------

wb = load_workbook(WB_PATH, data_only=True)


def sheet(name):
    if name not in wb.sheetnames:
        sys.exit(f"Missing sheet: {name} (found: {wb.sheetnames})")
    return wb[name]


# Reference: free-form key/value notes
reference = []
for k, v in sheet("Reference").iter_rows(values_only=True):
    k, v = clean(k), clean(v)
    if k is None and v is None:
        continue
    reference.append({"label": k, "value": v})

# Instruments
instruments = []
for r in rows_as_dicts(sheet("Instruments")):
    instruments.append({
        "name": r.get("Instrument"),
        "family": r.get("Family"),
        "patternBank": r.get("Pattern bank used"),
        "roles": r.get("Typical role(s)"),
        "rangeLow": r.get("Range low"),
        "rangeHigh": r.get("Range high"),
        "polyphony": r.get("Polyphony"),
        "timbre": r.get("Timbre (synth hint)"),
        "notes": r.get("Notes"),
    })

# Register map
register_map = []
for r in rows_as_dicts(sheet("Register_Map")):
    register_map.append({
        "role": r.get("Role"),
        "defaultRegister": r.get("Default register"),
        "positionVsMelody": r.get("Position vs melody"),
        "voicingRule": r.get("Voicing / collision rule"),
        "hardClamp": r.get("Hard clamp"),
    })

# Chord vocabulary
chord_vocab = []
for r in rows_as_dicts(sheet("Chord_Vocabulary")):
    chord_vocab.append({
        "symbol": r.get("Symbol (numeral form)"),
        "name": r.get("Name"),
        "chordNameInC": r.get("Chord-name form (key of C)"),
        "tones": r.get("Tones (vs root)"),
        "tier": r.get("Tier"),
    })
VOCAB_SHAPES = None  # set after helpers are defined

# Genre chord bank (genre cell forward-filled across its sections)
genres = {}
order = []
current = None
for r in rows_as_dicts(sheet("Genre_Chord_Bank")):
    g = r.get("Genre")
    if g:
        current = g
        if g not in genres:
            genres[g] = []
            order.append(g)
    if current is None:
        warn(f"Section row before any genre: {r}")
        continue
    bars = parse_bars(r.get("Bars"))
    chords = [c.strip() for c in str(r.get("Chords (1 per bar)") or "").split("-") if c.strip()]
    genres[current].append({
        "section": r.get("Section"),
        "bars": bars,
        "chords": chords,
    })

# Pattern banks
guitar_bank, piano_bank, bass_bank, ensemble_bank = [], [], [], []
for r in rows_as_dicts(sheet("Guitar_Strum_Bank")):
    guitar_bank.append(pattern_common(r))
for r in rows_as_dicts(sheet("Piano_Pattern_Bank")):
    p = pattern_common(r)
    p["hands"] = r.get("Hands")
    p["pedal"] = r.get("Pedal")
    piano_bank.append(p)
for r in rows_as_dicts(sheet("Bass_Pattern_Bank")):
    p = pattern_common(r)
    p["noteChoice"] = r.get("Note choice")
    bass_bank.append(p)
for r in rows_as_dicts(sheet("Ensemble_Pattern_Bank")):
    p = pattern_common(r)
    p["suits"] = r.get("Suits")
    ensemble_bank.append(p)

# Genre pattern map
pattern_map = []
for r in rows_as_dicts(sheet("Genre_Pattern_Map")):
    pattern_map.append({
        "genre": r.get("Genre"),
        "guitarPrimary": r.get("Primary accompaniment pattern"),
        "guitarAlternates": split_list(r.get("Alternate patterns")),
        "bassPattern": r.get("Bass pattern (Bass_Pattern_Bank)"),
        "defaultBpm": r.get("Default BPM"),
        "meter": r.get("Meter"),
        "notes": r.get("Notes"),
        "pianoPrimary": r.get("Piano primary pattern"),
        "pianoAlternates": split_list(r.get("Piano alternates")),
        "ensembleSuggestion": r.get("Ensemble suggestion (Ensemble_Pattern_Bank)"),
    })

# --- validation --------------------------------------------------------------

bank_genres = set(genres.keys())
map_genres = {m["genre"] for m in pattern_map}
for g in sorted(bank_genres - map_genres):
    warn(f"Genre in chord bank but not in pattern map: {g!r}")
for g in sorted(map_genres - bank_genres):
    warn(f"Genre in pattern map but not in chord bank: {g!r}")


def names(bank):
    return {p["name"] for p in bank}


guitar_names, piano_names = names(guitar_bank), names(piano_bank)
bass_names, ensemble_names = names(bass_bank), names(ensemble_bank)


def in_bank(val, bank):
    """Exact name match first; then with a trailing usage note '(...)' stripped,
    since bank names may themselves contain parens ('Block chords (half)')."""
    if val in bank:
        return True
    return re.sub(r"\s*\([^)]*\)\s*$", "", val).strip() in bank


# Chord templates: vocab exemplars like 'I', 'i7 / ii7', 'V65 / V43 / V42'
# stand for every scale degree of that shape. Normalize any numeral to its
# (case, suffix) shape and check that shape against the vocab.
ROMAN = re.compile(r"^[b#]?([ivxIVX]+)(.*)$")


def chord_shape(sym):
    m = ROMAN.match(sym.strip())
    if not m:
        return None
    numeral, suffix = m.group(1), m.group(2)
    case = "U" if numeral[0].isupper() else "L"
    return (case, suffix)


def vocab_shapes(vocab):
    shapes = set()
    for c in vocab:
        for form in str(c["symbol"] or "").split("/"):
            sh = chord_shape(form)
            if sh:
                shapes.add(sh)
    shapes.add(("U", "/V"))  # secondary dominant V/V parses oddly; allow it
    return shapes


VOCAB_SHAPES = vocab_shapes(chord_vocab)


for m in pattern_map:
    g = m["genre"]
    for label, val, bank in (
        ("guitar primary", m["guitarPrimary"], guitar_names),
        ("bass", m["bassPattern"], bass_names),
        ("piano primary", m["pianoPrimary"], piano_names),
    ):
        if val and not in_bank(val, bank):
            warn(f"{g}: {label} pattern not found in bank: {val!r}")
    for val in m["guitarAlternates"]:
        if not in_bank(val, guitar_names):
            warn(f"{g}: guitar alternate not in bank: {val!r}")
    for val in m["pianoAlternates"]:
        if not in_bank(val, piano_names):
            warn(f"{g}: piano alternate not in bank: {val!r}")
    es = m["ensembleSuggestion"]
    if es and es not in ("\u2014", "-"):
        # can be compound: 'A (strings) + B (horns)'
        for part in re.split(r"\s*\+\s*", es):
            if not in_bank(part.strip(), ensemble_names):
                warn(f"{g}: ensemble suggestion not in bank: {part.strip()!r}")

for g, sections in genres.items():
    prev_end = 0
    for s in sections:
        if s["bars"]:
            start, end = s["bars"]
            if start != prev_end + 1:
                warn(f"{g} / {s['section']}: bars {start}\u2013{end} not contiguous "
                     f"(previous section ended at {prev_end})")
            nbars = end - start + 1
            if len(s["chords"]) != nbars:
                warn(f"{g} / {s['section']}: {len(s['chords'])} chords for {nbars} bars")
            prev_end = end
        for c in s["chords"]:
            if chord_shape(c) not in VOCAB_SHAPES:
                warn(f"{g} / {s['section']}: chord {c!r} has no shape in Chord_Vocabulary")

# --- emit --------------------------------------------------------------------

data = {
    "meta": {
        "source": WB_PATH,
        "compiled": date.today().isoformat(),
        "referenceBpm": 60,
        "note": "All pattern slot durations are ms at 60 BPM; scale by 60/BPM at runtime.",
    },
    "reference": reference,
    "instruments": instruments,
    "registerMap": register_map,
    "chordVocabulary": chord_vocab,
    "genres": [{"name": g, "sections": genres[g]} for g in order],
    "patternBanks": {
        "guitar": guitar_bank,
        "piano": piano_bank,
        "bass": bass_bank,
        "ensemble": ensemble_bank,
    },
    "genrePatternMap": pattern_map,
}

js = ("// song_data.js \u2014 GENERATED by compile_data.py from " + WB_PATH + "\n"
      "// Do not hand-edit; edit the workbook and re-run the compiler.\n"
      "const SONG_DATA = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n")

with open(OUT_PATH, "w", encoding="utf-8") as f:
    f.write(js)

if EMBED:
    html = open(EMBED, encoding="utf-8").read()
    start_tag, end_tag = "<!-- SONG_DATA_BEGIN", "<!-- SONG_DATA_END -->"
    a, b = html.find(start_tag), html.find(end_tag)
    if a < 0 or b < 0:
        sys.exit(f"{EMBED}: SONG_DATA_BEGIN / SONG_DATA_END markers not found")
    a = html.find("\n", a) + 1
    html = html[:a] + "<script>\n" + js + "</script>\n" + html[b:]
    with open(EMBED, "w", encoding="utf-8", newline="\n") as f:
        f.write(html)
    print(f"Embedded into {EMBED}.")

n_pat = sum(len(b) for b in data["patternBanks"].values())
print(f"Wrote {OUT_PATH}: {len(genres)} genres, {n_pat} patterns, "
      f"{len(chord_vocab)} chord symbols, {len(instruments)} instruments.")
if warnings:
    print(f"\n{len(warnings)} validation warning(s):", file=sys.stderr)
    for w in warnings:
        print(f"  - {w}", file=sys.stderr)
else:
    print("No validation warnings.")
